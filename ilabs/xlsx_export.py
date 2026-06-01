"""
Append a training record row to a Training Schedule xlsx file.

The workbook is expected to have a header row (row 1).  Known headers are
mapped to record fields; form-data columns are matched by key name
(case-insensitive).  If the file doesn't exist yet it is created with a
default header row.

Requires:  pip install openpyxl
"""

import json
from pathlib import Path

try:
    import openpyxl
    import openpyxl.styles
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False


# Maps normalised xlsx header text → record field key
HEADER_MAP: dict[str, str] = {
    "date":               "training_date",
    "training date":      "training_date",
    "trainingdate":       "training_date",
    "day":                "training_day",
    "training day":       "training_day",
    "trainingday":        "training_day",
    "time":               "training_time",
    "training time":      "training_time",
    "trainingtime":       "training_time",
    "microscope":         "microscope",
    "system":             "microscope",
    "instrument":         "microscope",
    "scope":              "microscope",
    "trainee":            "owner_name",
    "trainee name":       "owner_name",
    "traineename":        "owner_name",
    "requester":          "owner_name",
    "name":               "owner_name",
    "first name":         "owner_name",   # best-effort
    "email":              "owner_email",
    "trainee email":      "owner_email",
    "traineeemail":       "owner_email",
    "pi":                 "pi_name",
    "pi name":            "pi_name",
    "piname":             "pi_name",
    "principal investigator": "pi_name",
    "pi email":           "pi_email",
    "piemail":            "pi_email",
    "core":               "core_lab",
    "core lab":           "core_lab",
    "corelab":            "core_lab",
    "notes":              "local_notes",
    "request id":         "request_id",
    "requestid":          "request_id",
    "id":                 "request_id",
    "service":            "service_name",
    "request name":       "name",
    "requestname":        "name",
    "title":              "name",
    "class":              "class_taken",
    "class taken":        "class_taken",
}

# Column order used when creating a brand-new xlsx (or when row 1 is empty)
DEFAULT_HEADERS = [
    "Date", "Day", "Time", "Microscope",
    "Trainee Name", "Email", "PI Name", "PI Email",
    "Core", "Service", "Request ID", "Class", "Notes",
]
DEFAULT_FIELDS = [
    "training_date", "training_day", "training_time", "microscope",
    "owner_name",    "owner_email",  "pi_name",       "pi_email",
    "core_lab",      "service_name", "request_id",    "class_taken", "local_notes",
]


def _norm(text: str) -> str:
    """Lowercase + collapse whitespace."""
    return " ".join(text.lower().split())


def append_training_row(rec: dict, xlsx_path: str) -> None:
    """
    Append one training record row to the workbook at *xlsx_path*.

    - If the file does not exist it is created with DEFAULT_HEADERS.
    - If row 1 is empty, DEFAULT_HEADERS are written first.
    - Otherwise the existing headers determine column placement.

    Raises:
        ImportError   if openpyxl is not installed.
        Exception     propagated from openpyxl on file errors.
    """
    if not HAS_OPENPYXL:
        raise ImportError(
            "openpyxl is required for xlsx export.\n"
            "Install it with:  pip install openpyxl"
        )

    form_data: dict = json.loads(rec.get("form_data") or "{}")
    p = Path(xlsx_path)

    if p.exists():
        wb = openpyxl.load_workbook(xlsx_path)
        ws = wb.active
        headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
        headers = [str(h).strip() if h is not None else "" for h in headers]
        if all(h == "" for h in headers):
            # Blank first row — write default headers
            _write_headers(ws, DEFAULT_HEADERS)
            headers = DEFAULT_HEADERS
    else:
        # Create new workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Training Schedule"
        _write_headers(ws, DEFAULT_HEADERS)
        headers = DEFAULT_HEADERS

    # ── Build row matching header order ───────────────────────────────────────
    row: list = []
    for header in headers:
        if not header:
            row.append("")
            continue
        n = _norm(header)
        field = HEADER_MAP.get(n)
        if field:
            val = rec.get(field, "") or ""
            if field == "class_taken":
                val = "Yes" if val == "1" else ""
            row.append(val)
        else:
            # Try matching a form_data key (exact, then case-insensitive)
            fd_val = (
                form_data.get(header)
                or form_data.get(header.title())
                or next((v for k, v in form_data.items()
                         if k.lower() == header.lower()), "")
            )
            row.append(fd_val or "")

    ws.append(row)
    wb.save(xlsx_path)


def _write_headers(ws, headers: list) -> None:
    bold = openpyxl.styles.Font(bold=True)
    for c, h in enumerate(headers, start=1):
        cell = ws.cell(1, c)
        cell.value = h
        cell.font  = bold
