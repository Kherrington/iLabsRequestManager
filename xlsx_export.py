"""
Append a training record row to a Training Schedule xlsx file.

The workbook is expected to have a header row (row 1).  Known headers are
mapped to record fields; form-data columns are matched by key name
(case-insensitive).  If the file does not exist it is created with a
default header row.

Requires:  pip install openpyxl
"""

import json
from pathlib import Path

try:
    import openpyxl
    import openpyxl.styles
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False


# Maps normalised xlsx header text → internal record field key.
# Add entries here whenever a new column name is encountered.
HEADER_MAP: dict[str, str] = {
    # ── Date / time ───────────────────────────────────────────────────────────
    "date":               "training_date",
    "training date":      "training_date",
    "trainingdate":       "training_date",
    "trng date":          "training_date",
    "day":                "training_day",
    "training day":       "training_day",
    "trainingday":        "training_day",
    "time":               "training_time",
    "training time":      "training_time",
    "trainingtime":       "training_time",
    # ── Microscope / system ───────────────────────────────────────────────────
    "microscope":         "microscope",
    "system":             "microscope",
    "instrument":         "microscope",
    "scope":              "microscope",
    # ── Requester / trainee ───────────────────────────────────────────────────
    "people":             "owner_name",   # common header in CALM/CVRI schedules
    "person":             "owner_name",
    "trainee":            "owner_name",
    "trainee name":       "owner_name",
    "traineename":        "owner_name",
    "requester":          "owner_name",
    "name":               "owner_name",
    "first name":         "owner_name",
    "email":              "owner_email",
    "trainee email":      "owner_email",
    "traineeemail":       "owner_email",
    # ── PI / lab ──────────────────────────────────────────────────────────────
    "pi":                 "pi_name",
    "pi name":            "pi_name",
    "piname":             "pi_name",
    "lab":                "pi_name",
    "lab name":           "pi_name",
    "principal investigator": "pi_name",
    "pi email":           "pi_email",
    "piemail":            "pi_email",
    # ── Core ─────────────────────────────────────────────────────────────────
    "core":               "core_lab",
    "core lab":           "core_lab",
    "corelab":            "core_lab",
    "facility":           "core_lab",
    # ── Class taken ───────────────────────────────────────────────────────────
    "class":              "class_taken",
    "class taken":        "class_taken",
    "classtaken":         "class_taken",
    # ── Notes ─────────────────────────────────────────────────────────────────
    "notes":              "local_notes",
    "note":               "local_notes",
    "comments":           "local_notes",
    # ── Admin ─────────────────────────────────────────────────────────────────
    "request id":         "request_id",
    "requestid":          "request_id",
    "id":                 "request_id",
    "service":            "service_name",
    "request name":       "name",
    "requestname":        "name",
    "title":              "name",
}

# Column order used when creating a brand-new xlsx (or when row 1 is empty)
DEFAULT_HEADERS = [
    "Date", "Day", "Time", "Microscope",
    "People", "Email", "Lab", "PI Email",
    "Core", "Class Taken", "Notes", "Request ID",
]
DEFAULT_FIELDS = [
    "training_date", "training_day", "training_time", "microscope",
    "owner_name",    "owner_email",  "pi_name",       "pi_email",
    "core_lab",      "class_taken",  "local_notes",   "request_id",
]


def _norm(text: str) -> str:
    """Lowercase + collapse whitespace."""
    return " ".join(text.lower().split())


def append_training_row(rec: dict, xlsx_path: str,
                        sheet_name: str = "") -> None:
    """
    Append one training record row to the workbook at *xlsx_path*.

    sheet_name  Name of the worksheet to append to.  If blank or not found,
                the active (first) sheet is used.

    - If the file does not exist it is created with DEFAULT_HEADERS.
    - If row 1 is empty, DEFAULT_HEADERS are written first.
    - Otherwise the existing headers determine column placement.

    Field values applied:
        People / Requester  → owner_name  (requester name)
        Microscope          → microscope
        Date                → training_date
        Day                 → training_day
        Time                → training_time
        Notes               → local_notes
        Class / Class Taken → "Yes" if class_taken == "1", else ""

    Raises:
        ImportError   if openpyxl is not installed.
        Exception     propagated from openpyxl on file errors.
    """
    if not HAS_OPENPYXL:
        raise ImportError(
            "openpyxl is required for xlsx export.\n"
            "Install it with:  pip install openpyxl"
        )

    form_data: dict = json.loads(rec.get("form_data") or "{}")
    p = Path(xlsx_path)

    # ── Open or create workbook ───────────────────────────────────────────────
    if p.exists():
        wb = openpyxl.load_workbook(xlsx_path)
        # Select sheet by name, fall back to active
        if sheet_name and sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
        else:
            ws = wb.active

        # Read headers from row 1, trimming phantom empty columns at the right
        # (formatted Excel files often have max_column >> actual data columns)
        raw_headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
        headers = [str(h).strip() if h is not None else "" for h in raw_headers]
        # Drop trailing empty headers
        while headers and not headers[-1]:
            headers.pop()

        if not headers:
            _write_headers(ws, DEFAULT_HEADERS)
            headers = DEFAULT_HEADERS
    else:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = sheet_name or "Training Schedule"
        _write_headers(ws, DEFAULT_HEADERS)
        headers = DEFAULT_HEADERS

    # ── Build row matching header order ───────────────────────────────────────
    row: list = []
    for header in headers:
        if not header:
            row.append("")
            continue
        n     = _norm(header)
        field = HEADER_MAP.get(n)
        if field:
            val = rec.get(field, "") or ""
            # class_taken: store "Yes" / "" rather than "1" / "0"
            if field == "class_taken":
                val = "Yes" if val == "1" else ""
            row.append(val)
        else:
            # Unknown column — try matching a form_data key
            fd_val = (
                form_data.get(header)
                or form_data.get(header.title())
                or next((v for k, v in form_data.items()
                         if k.lower() == header.lower()), "")
            )
            row.append(fd_val or "")

    ws.append(row)
    wb.save(xlsx_path)

    # Return a summary for debugging / status messages
    mapped   = {h: v for h, v in zip(headers, row) if v not in (None, "")}
    unmapped = [h for h, v in zip(headers, row) if v in (None, "")]
    return {"headers": headers, "written": mapped, "empty": unmapped}


def _write_headers(ws, headers: list) -> None:
    bold = openpyxl.styles.Font(bold=True)
    for c, h in enumerate(headers, start=1):
        cell = ws.cell(1, c)
        cell.value = h
        cell.font  = bold
